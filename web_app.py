from flask import Flask, render_template
from openpyxl import load_workbook
from pathlib import Path

base_dir = Path(__file__).resolve().parent
template_dir = base_dir / "templates"
static_dir = base_dir / "static"

app = Flask(
    __name__,
    template_folder=str(template_dir),
    static_folder=str(static_dir),
)


def load_first_customer() -> dict | None:
    try:
        excel_path = Path("ThongTinKhachHang.xlsx")
        if not excel_path.exists():
            return None
        wb = load_workbook(filename=str(excel_path))
        sheet = wb.active
        header = [cell.value for cell in sheet[1]]
        # Find first non-empty data row
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and any(cell is not None for cell in row):
                return dict(zip(header, row))
        return None
    except Exception:
        return None


def load_all_customers_for_web() -> list:
    """Load tất cả khách hàng từ Excel và phân nhóm"""
    try:
        excel_path = Path("ThongTinKhachHang.xlsx")
        if not excel_path.exists():
            return []
        
        wb = load_workbook(filename=str(excel_path))
        sheet = wb.active
        header = [cell.value for cell in sheet[1]]
        
        customers = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and any(cell is not None for cell in row):
                raw_data = dict(zip(header, row))
                
                # Chuẩn hóa dữ liệu
                customer = {
                    "code": raw_data.get("Mã KH") or raw_data.get("Ma KH") or raw_data.get("Code") or "",
                    "name": raw_data.get("Họ Tên") or raw_data.get("Ho Ten") or raw_data.get("Ten") or "",
                    "phone": raw_data.get("Số ĐT") or raw_data.get("So DT") or raw_data.get("SĐT") or "",
                    "email": raw_data.get("Email") or "",
                    "address": raw_data.get("Địa Chỉ") or raw_data.get("Dia Chi") or "",
                    "total_amount": raw_data.get("Tổng tiền mua") or raw_data.get("Tong Tien Mua") or "0đ",
                    "last_purchase": raw_data.get("Ngày cuối mua") or raw_data.get("Ngay Cuoi Mua") or "",
                }
                
                # Phân nhóm dựa trên tổng tiền mua
                total_amount_str = str(customer["total_amount"]).replace("đ", "").replace(",", "")
                try:
                    total_amount = float(total_amount_str) if total_amount_str else 0
                except ValueError:
                    total_amount = 0
                
                if total_amount >= 10000000:  # >= 10 triệu
                    customer["group"] = "vip"
                elif total_amount >= 5000000:  # >= 5 triệu
                    customer["group"] = "loyal"
                else:
                    customer["group"] = "potential"
                
                # Phân trạng thái Active/Inactive dựa trên lần cuối mua hàng
                last_purchase = customer["last_purchase"]
                if last_purchase and last_purchase != "Chưa có":
                    # Giả sử nếu có mua hàng trong 6 tháng gần đây thì Active
                    customer["status"] = "active"
                else:
                    customer["status"] = "inactive"
                
                customers.append(customer)
        
        # Sắp xếp theo alphabet dựa trên chữ cái cuối của tên
        customers.sort(key=lambda x: x["name"][-1].lower() if x["name"] else "z")
        
        return customers
    except Exception:
        return []


def get_customer_stats(customers: list) -> dict:
    """Tính toán thống kê khách hàng"""
    active_customers = len([c for c in customers if c["status"] == "active"])
    
    # Tính tổng tiền của khách hàng Active
    active_amount = 0
    for customer in customers:
        if customer["status"] == "active":
            total_amount_str = str(customer["total_amount"]).replace("đ", "").replace(",", "")
            try:
                active_amount += float(total_amount_str) if total_amount_str else 0
            except ValueError:
                pass
    
    # Giả sử số đơn hàng hoạt động = số khách hàng active * 2 (trung bình)
    active_orders = active_customers * 2
    
    stats = {
        "total_customers": len(customers),
        "active_customers": active_customers,
        "active_orders": active_orders,
        "active_amount": f"{active_amount:,.0f}đ" if active_amount > 0 else "0đ",
    }
    return stats


@app.route("/")
def home():
    return render_template("base.html", active="overview")


@app.route("/customer-dashboard")
def customer_dashboard():
    from flask import request
    customer_code = request.args.get('code')
    
    # Load khách hàng theo code nếu có,否则 load khách hàng đầu tiên
    if customer_code:
        customers = load_all_customers_for_web()
        raw = next((c for c in customers if c.get("code") == customer_code), None)
    else:
        raw = load_first_customer()
    
    customer = None
    if raw:
        customer = {
            "name": raw.get("name") or raw.get("Họ Tên") or raw.get("Ho Ten") or raw.get("Ten") or "",
            "dob": raw.get("Ngày sinh") or raw.get("Ngay Sinh") or "",
            "phone": raw.get("phone") or raw.get("Số ĐT") or raw.get("So DT") or raw.get("SĐT") or raw.get("Phone") or "",
            "email": raw.get("email") or raw.get("Email") or "",
            "address": raw.get("address") or raw.get("Địa Chỉ") or raw.get("Dia Chi") or raw.get("Address") or "",
            "code": raw.get("code") or raw.get("Mã KH") or raw.get("Ma KH") or raw.get("Code") or "",
            "total": raw.get("total_amount") or raw.get("Tổng tiền mua") or raw.get("Tong Tien Mua") or "",
            "last_purchase": raw.get("last_purchase") or raw.get("Ngày cuối mua") or raw.get("Ngay Cuoi Mua") or "",
        }

    sample_orders = [
        {
            "code": "DLU00001",
            "customer": (customer["name"] if customer and customer.get("name") else "Nguyễn Phước Lộc"),
            "export_status": "Đã xuất",
            "value": "1,500,000đ",
            "date": "28/09/2025",
            "status": "done",
            "status_label": "Hoàn thành",
        },
        {
            "code": "DLU00002",
            "customer": (customer["name"] if customer and customer.get("name") else "Nguyễn Phước Lộc"),
            "export_status": "Chưa xuất",
            "value": "2,000,000đ",
            "date": "29/09/2025",
            "status": "processing",
            "status_label": "Đang xử lý",
        },
    ]

    return render_template(
        "customer_dashboard_jinja.html",
        active="customers",
        user_name="Admin",
        customer=customer,
        orders=sample_orders,
    )


@app.route("/customers_list")
@app.route("/customers")
def customers_list():
    customers = load_all_customers_for_web()
    stats = get_customer_stats(customers)
    
    return render_template(
        "customer_list.html",
        active="customers",
        customers=customers,
        stats=stats,
    )


if __name__ == "__main__":
    app.run(debug=True)


